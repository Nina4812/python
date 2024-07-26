

# Programmer: Your Name
# Title: Sysbench Results Parser and Excel Writer
# Dependencies: re, openpyxl, os
# Description: This script parses Sysbench results from a text file and writes the parsed data into an Excel file,
# creating separate tables for CPU, Memory, FileIO (all modes), and System Info. It supports both creating a new file
# and appending to an existing file. The script also formats the tables for better readability.


import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import os

def ParseSysbenchResults(file_path):
    """Parse the sysbench results from the given file."""
    with open(file_path, 'r') as file:
        content = file.read()

    sections = content.split('---------------------')

    def ExtractCpuInfo(section):
        """Extract CPU information from the given section."""
        cpu_info = {}
        cpu_info['Command'] = re.search(r'Command: (.*)', section)
        cpu_info['Events_per_second'] = re.search(r'events per second:\s+([\d.]+)', section)
        cpu_info['Total_time'] = re.search(r'total time:\s+([\d.]+)s', section)
        cpu_info['Total_number_of_events'] = re.search(r'total number of events:\s+([\d.]+)', section)
        cpu_info['Latency_min'] = re.search(r'min:\s+([\d.]+)', section)
        cpu_info['Latency_avg'] = re.search(r'avg:\s+([\d.]+)', section)
        cpu_info['Latency_max'] = re.search(r'max:\s+([\d.]+)', section)
        cpu_info['Latency_95th_percentile'] = re.search(r'95th percentile:\s+([\d.]+)', section)
        cpu_info['Latency_sum'] = re.search(r'sum:\s+([\d.]+)', section)
        cpu_info['Events_avg'] = re.search(r'events \(avg/stddev\):\s+([\d.]+)/', section)
        cpu_info['Execution_time_avg'] = re.search(r'execution time \(avg/stddev\):\s+([\d.]+)/', section)

        return {k: v.group(1) if v else 'N/A' for k, v in cpu_info.items()}

    def ExtractMemoryInfo(section):
        """Extract Memory information from the given section."""
        memory_info = {}
        memory_info['Command'] = re.search(r'Command: (.*)', section)
        memory_info['Total_operations'] = re.search(r'Total operations:\s+([\d.]+)', section)
        memory_info['Operations_per_second'] = re.search(r'\( *([\d.]+) per second\)', section)
        memory_info['MiB_transferred'] = re.search(r'([\d.]+) MiB transferred', section)
        memory_info['MiB_per_sec'] = re.search(r'MiB transferred \(([\d.]+)', section)
        memory_info['Total_time'] = re.search(r'total time:\s+([\d.]+)s', section)
        memory_info['Total_number_of_events'] = re.search(r'total number of events:\s+([\d.]+)', section)
        memory_info['Latency_min'] = re.search(r'min:\s+([\d.]+)', section)
        memory_info['Latency_avg'] = re.search(r'avg:\s+([\d.]+)', section)
        memory_info['Latency_max'] = re.search(r'max:\s+([\d.]+)', section)
        memory_info['Latency_95th_percentile'] = re.search(r'95th percentile:\s+([\d.]+)', section)
        memory_info['Latency_sum'] = re.search(r'sum:\s+([\d.]+)', section)
        memory_info['Events_avg'] = re.search(r'events \(avg/stddev\):\s+([\d.]+)/', section)
        memory_info['Execution_time_avg'] = re.search(r'execution time \(avg/stddev\):\s+([\d.]+)/', section)

        return {k: v.group(1) if v else 'N/A' for k, v in memory_info.items()}

    def ExtractFileioInfo(section):
        """Extract FileIO information from the given section."""
        fileio_info = {}
        fileio_info['Command'] = re.search(r'Command: (.*)', section)
        fileio_info['Reads_per_second'] = re.search(r'reads/s:\s+([\d.]+)', section)
        fileio_info['Writes_per_second'] = re.search(r'writes/s:\s+([\d.]+)', section)
        fileio_info['Fsyncs_per_second'] = re.search(r'fsyncs/s:\s+([\d.]+)', section)
        fileio_info['Throughput_read'] = re.search(r'read, MiB/s:\s+([\d.]+)', section)
        fileio_info['Throughput_write'] = re.search(r'written, MiB/s:\s+([\d.]+)', section)
        fileio_info['Total_time'] = re.search(r'total time:\s+([\d.]+)s', section)
        fileio_info['Total_number_of_events'] = re.search(r'total number of events:\s+([\d.]+)', section)
        fileio_info['Latency_min'] = re.search(r'min:\s+([\d.]+)', section)
        fileio_info['Latency_avg'] = re.search(r'avg:\s+([\d.]+)', section)
        fileio_info['Latency_max'] = re.search(r'max:\s+([\d.]+)', section)
        fileio_info['Latency_95th_percentile'] = re.search(r'95th percentile:\s+([\d.]+)', section)
        fileio_info['Latency_sum'] = re.search(r'sum:\s+([\d.]+)', section)
        fileio_info['Events_avg'] = re.search(r'events \(avg/stddev\):\s+([\d.]+)/', section)
        fileio_info['Execution_time_avg'] = re.search(r'execution time \(avg/stddev\):\s+([\d.]+)/', section)

        return {k: v.group(1) if v else 'N/A' for k, v in fileio_info.items()}

    def ExtractAdditionalInfo(section):
        """Extract additional system information from the given section."""
        additional_info = {}
        additional_info['Date_Time'] = re.search(r'Date and Time:\s+(.*)', section)
        additional_info['Host_Name'] = re.search(r'Host Name:\s+(.*)', section)
        additional_info['System_Architecture'] = re.search(r'System Architecture:\s+(.*)', section)
        additional_info['Kernel_Version'] = re.search(r'Kernel Version:\s+(.*)', section)
        additional_info['Distribution_Info'] = re.search(r'Distribution Info:\s+(.*)', section)
        additional_info['CPU_Info'] = re.search(r'CPU Info:\s+(.*)', section)
        additional_info['Total_Memory'] = re.search(r'Total Memory:\s+(.*)', section)
        additional_info['Total_Swap_Space'] = re.search(r'Total Swap Space:\s+(.*)', section)
        return {k: v.group(1) if v else 'N/A' for k, v in additional_info.items()}

    results = {
        "Additional_Info": ExtractAdditionalInfo(sections[0]),
        "CPU": ExtractCpuInfo(sections[1]),
        "Memory": ExtractMemoryInfo(sections[2]),
        "FileIO_seqwr": ExtractFileioInfo(sections[3]),
        "FileIO_seqrewr": ExtractFileioInfo(sections[4]),
        "FileIO_seqrd": ExtractFileioInfo(sections[5]),
        "FileIO_rndrd": ExtractFileioInfo(sections[6]),
        "FileIO_rndwr": ExtractFileioInfo(sections[7]),
        "FileIO_rndrw": ExtractFileioInfo(sections[8])
    }

    return results

def WriteToExcel(data, excel_path, mode):
    """Write the parsed data to an Excel file in either 'new' or 'append' mode."""
    def write_data(sheet, start_row, data_dict, title):
        """Helper function to write a section of data to the sheet."""
        headers = ["Metric", title]
        sheet.append(headers)
        for cell in sheet[start_row]:
            cell.font = Font(bold=True)

        metrics = set(data_dict.keys())
        colored_metrics = {
            "CPU": ["Total_time", "Total_number_of_events", "Latency_min", "Latency_avg", "Latency_max",
                    "Latency_95th_percentile", "Latency_sum"],
            "Memory": ["Total_operations", "Operations_per_second", "MiB_transferred", "MiB_per_sec", "Total_time",
                       "Total_number_of_events", "Latency_min", "Latency_avg", "Latency_max",
                       "Latency_95th_percentile", "Latency_sum"],
            "FileIO_seqwr": ["Throughput_read", "Throughput_write", "Latency_min", "Latency_avg", "Latency_max",
                             "Latency_95th_percentile", "Latency_sum"],
            "FileIO_seqrewr": ["Throughput_read", "Throughput_write", "Latency_min", "Latency_avg", "Latency_max",
                               "Latency_95th_percentile", "Latency_sum"],
            "FileIO_seqrd": ["Throughput_read", "Throughput_write", "Latency_min", "Latency_avg", "Latency_max",
                             "Latency_95th_percentile", "Latency_sum"],
            "FileIO_rndrd": ["Throughput_read", "Throughput_write", "Latency_min", "Latency_avg", "Latency_max",
                             "Latency_95th_percentile", "Latency_sum"],
            "FileIO_rndwr": ["Throughput_read", "Throughput_write", "Latency_min", "Latency_avg", "Latency_max",
                             "Latency_95th_percentile", "Latency_sum"],
            "FileIO_rndrw": ["Throughput_read", "Throughput_write", "Latency_min", "Latency_avg", "Latency_max",
                             "Latency_95th_percentile", "Latency_sum"],
        }

        color_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

        for metric in sorted(metrics):
            row = [metric, data_dict.get(metric, 'N/A')]
            sheet.append(row)

        for cell in sheet[start_row + 1]:
            cell.alignment = Alignment(wrap_text=True)

        for cell in sheet['A']:
            cell.font = Font(bold=True)

        for row in sheet.iter_rows(min_row=start_row + 1, max_col=len(headers), max_row=sheet.max_row):
            metric = row[0].value
            if metric in colored_metrics.get(title, []):
                for cell in row:
                    cell.fill = color_fill
                    cell.font = Font(bold=True)

        # Adjust column width
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

    if mode == 'new' or not os.path.exists(excel_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sysbench Results"

        # Write each section
        write_data(sheet, 1, data["CPU"], "CPU")
        write_data(sheet, sheet.max_row +1, data["Memory"], "Memory")
        write_data(sheet, sheet.max_row +1, data["FileIO_seqwr"], "FileIO_seqwr")
        write_data(sheet, sheet.max_row+1 , data["FileIO_seqrewr"], "FileIO_seqrewr")
        write_data(sheet, sheet.max_row+1 , data["FileIO_seqrd"], "FileIO_seqrd")
        write_data(sheet, sheet.max_row+1 , data["FileIO_rndrd"], "FileIO_rndrd")
        write_data(sheet, sheet.max_row +1, data["FileIO_rndwr"], "FileIO_rndwr")
        write_data(sheet, sheet.max_row +1, data["FileIO_rndrw"], "FileIO_rndrw")

        # Add additional system information at the end
        additional_start_row = sheet.max_row + 2
        sheet.append([])  # Add an empty row for separation
        sheet.append(["Additional Info"])
        for key, value in data["Additional_Info"].items():
            sheet.append([key, value])

        workbook.save(excel_path)
        print(f"The results are saved in {excel_path}")
    elif mode == 'append':
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active

        start_row = sheet.max_row + 2

        # Write each section
        write_data(sheet, start_row, data["CPU"], "CPU")
        write_data(sheet, sheet.max_row+1 , data["Memory"], "Memory")
        write_data(sheet, sheet.max_row +1, data["FileIO_seqwr"], "FileIO_seqwr")
        write_data(sheet, sheet.max_row +1, data["FileIO_seqrewr"], "FileIO_seqrewr")
        write_data(sheet, sheet.max_row +1, data["FileIO_seqrd"], "FileIO_seqrd")
        write_data(sheet, sheet.max_row+1, data["FileIO_rndrd"], "FileIO_rndrd")
        write_data(sheet, sheet.max_row +1, data["FileIO_rndwr"], "FileIO_rndwr")
        write_data(sheet, sheet.max_row+1 , data["FileIO_rndrw"], "FileIO_rndrw")

        # Add additional system information at the end
        additional_start_row = sheet.max_row + 2
        sheet.append([])  # Add an empty row for separation
        sheet.append(["Additional Info"])
        for key, value in data["Additional_Info"].items():
            sheet.append([key, value])

        workbook.save(excel_path)
        print(f"The results are appended to {excel_path}")

def main():
    text_file_path = input("Enter the path of the Sysbench results text file: ")
    excel_file_path = input("Enter the desired path for the output Excel file: ")
    mode = input("Enter 'new' to create a new Excel file or 'append' to update the existing Excel file: ").strip().lower()

    if mode not in ['new', 'append']:
        print("Invalid mode selected. Please enter 'new' or 'append'.")
        return

    sysbench_results = ParseSysbenchResults(text_file_path)
    WriteToExcel(sysbench_results, excel_file_path, mode)

if __name__ == "__main__":
    main()
